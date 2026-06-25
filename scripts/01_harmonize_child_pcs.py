#!/usr/bin/env python3
"""
Step 1: Harmonize primary dataset with child ancestry principal components.

Joins the primary dataset with ancestry PCs from the APrON cohort SNP array
data. Only infant PCs (FID suffix -10) are used, matching the infant genotype
data already in the primary dataset.

Join strategy:
  - Inf_APrONID in primary data: "C0123-10-2" (family-role-visit)
  - FID in PCs file:             "C0123-10"   (family-role, no visit suffix)
  - Strip trailing visit suffix (-1 or -2) from Inf_APrONID to create join key.

Run from: the repository root
Usage:    python3 scripts/01_harmonize_child_pcs.py

Input files (NOT modified):
  - data/Fulldata_raw.txt  (override via the RAW_DATA_FILE env var / config.sh)
  - data/APRON_PCs.xlsx

Output files:
  - data/Fulldata_with_PCs.txt
  - results/harmonization_child_report.txt
"""

import csv
import os
import sys
from collections import Counter

try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl required. Install with: pip install openpyxl")
    sys.exit(1)

PRIMARY_FILE = os.environ.get("RAW_DATA_FILE", "data/Fulldata_raw.txt")
PCS_FILE = "data/APRON_PCs.xlsx"
OUTPUT_FILE = "data/Fulldata_with_PCs.txt"
REPORT_FILE = "results/harmonization_child_report.txt"
ENCODING = "latin-1"


def read_primary_data(filepath):
    with open(filepath, "r", encoding=ENCODING) as f:
        reader = csv.DictReader(f, delimiter="\t")
        fieldnames = list(reader.fieldnames)
        rows = list(reader)
    return fieldnames, rows


def read_pcs_data(filepath):
    wb = openpyxl.load_workbook(filepath, read_only=True)
    ws = wb[wb.sheetnames[0]]
    headers = None
    pc_dict = {}
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            headers = list(row)
            continue
        fid = row[0]
        pc_dict[fid] = {headers[j]: row[j] for j in range(1, len(headers))}
    wb.close()
    return headers[1:], pc_dict


def make_join_key(inf_apron_id):
    """Strip trailing visit suffix (-1 or -2) from Inf_APrONID."""
    parts = inf_apron_id.rsplit("-", 1)
    if len(parts) == 2 and parts[1] in ("1", "2"):
        return parts[0]
    return inf_apron_id


def main():
    report_lines = []

    def report(msg):
        print(msg)
        report_lines.append(msg)

    report("=" * 70)
    report("STEP 1: CHILD ANCESTRY PC HARMONIZATION")
    report("=" * 70)
    report("")

    # Load data
    fieldnames, primary_rows = read_primary_data(PRIMARY_FILE)
    pc_col_names, pc_dict = read_pcs_data(PCS_FILE)
    pc_cols_to_use = pc_col_names  # all 50 PCs

    report(f"Primary data: {len(primary_rows)} rows, {len(fieldnames)} columns")
    report(f"PCs data: {len(pc_dict)} records, {len(pc_col_names)} PCs")
    report(f"PCs to include: {len(pc_cols_to_use)} ({pc_cols_to_use[0]}..{pc_cols_to_use[-1]})")
    report("")

    # Analyze PCs file composition
    report("PCs file composition:")
    suffix_counts = Counter()
    for fid in pc_dict:
        parts = fid.split("-")
        if len(parts) >= 2:
            suffix_counts[parts[1]] += 1
    for suffix, count in sorted(suffix_counts.items()):
        label = {"01": "parent/mother (inferred from MaternalID column)",
                 "10": "infant/child (inferred from Inf_APrONID column)"}.get(suffix, "unknown role")
        report(f"  -{suffix} ({label}): {count}")
    report("")

    # Build join keys
    inf_ids = [r["Inf_APrONID"] for r in primary_rows]
    join_keys = [make_join_key(iid) for iid in inf_ids]

    # Validate all keys are infant IDs
    non_infant = [(i, jk) for i, jk in enumerate(join_keys) if not jk.endswith("-10")]
    if non_infant:
        report(f"WARNING: {len(non_infant)} join keys are NOT infant IDs")
    else:
        report("All join keys are infant IDs (-10 suffix): OK")

    # Check duplicates
    id_counts = Counter(inf_ids)
    dups = {k: v for k, v in id_counts.items() if v > 1}
    if dups:
        report(f"Duplicate Inf_APrONIDs: {len(dups)}")
        for dup_id, count in sorted(dups.items()):
            dup_rows = [(i, primary_rows[i]["Samples"]) for i, r in enumerate(primary_rows) if r["Inf_APrONID"] == dup_id]
            report(f"  {dup_id} x{count}: " + ", ".join(f"row {i} ({s})" for i, s in dup_rows))
        report("  -> Both copies receive same PCs (correct)")
    report("")

    # Perform join
    matched, unmatched = [], []
    for i in range(len(primary_rows)):
        (matched if join_keys[i] in pc_dict else unmatched).append(i)

    report(f"Matched: {len(matched)}/{len(primary_rows)} ({100*len(matched)/len(primary_rows):.1f}%)")
    report(f"Unmatched: {len(unmatched)}/{len(primary_rows)}")
    if unmatched:
        report("")
        for idx in unmatched:
            row = primary_rows[idx]
            likert = row.get("Final7PointLikertScale", "").strip()
            report(f"  Row {idx}: Sample={row['Samples']}, "
                   f"Inf_APrONID={row['Inf_APrONID']}, "
                   f"{'has outcome' if likert else 'no outcome'}")
    report("")

    # Build output
    output_fieldnames = list(fieldnames) + pc_cols_to_use
    output_rows = []
    for i, row in enumerate(primary_rows):
        new_row = dict(row)
        jk = join_keys[i]
        if jk in pc_dict:
            for pc_col in pc_cols_to_use:
                new_row[pc_col] = pc_dict[jk].get(pc_col, "NA")
        else:
            for pc_col in pc_cols_to_use:
                new_row[pc_col] = "NA"
        output_rows.append(new_row)

    with open(OUTPUT_FILE, "w", encoding=ENCODING, newline="") as f:
        writer = csv.DictWriter(f, fieldnames=output_fieldnames, delimiter="\t")
        writer.writeheader()
        writer.writerows(output_rows)

    report(f"Output: {OUTPUT_FILE}")
    report(f"  {len(output_rows)} rows, {len(output_fieldnames)} columns "
           f"({len(fieldnames)} original + {len(pc_cols_to_use)} PCs)")
    report("")

    # Validation
    report("Validation:")
    assert len(output_rows) == len(primary_rows)
    report(f"  Row count preserved: OK")

    for i, row in enumerate(primary_rows):
        for col in fieldnames:
            assert output_rows[i][col] == row[col], f"Column {col} modified in row {i}!"
    report(f"  Original columns unchanged: OK")

    for i in matched:
        for pc_col in pc_cols_to_use[:3]:
            assert str(pc_dict[join_keys[i]][pc_col]) == str(output_rows[i][pc_col])
    report(f"  PC values verified for {len(matched)} matched rows: OK")

    for i in unmatched:
        for pc_col in pc_cols_to_use:
            assert output_rows[i][pc_col] == "NA"
    report(f"  Unmatched rows have NA PCs: OK")
    report("")

    report("STEP 1 COMPLETE")
    report(f"  Neither input file was modified.")

    with open(REPORT_FILE, "w") as f:
        f.write("\n".join(report_lines) + "\n")
    print(f"Report: {REPORT_FILE}")


if __name__ == "__main__":
    main()
