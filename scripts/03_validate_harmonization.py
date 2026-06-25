#!/usr/bin/env python3
"""
Step 3: Independent validation of the fully harmonized dataset.

Reads all source files from scratch and performs exhaustive cross-checks
against the final output (Fulldata_with_PCs_and_maternal_PCs.txt) to ensure
no annotation errors. This script does NOT modify any files.

Checks performed:
  1. Row count matches primary source
  2. All 59 original columns unchanged vs primary source (cell-by-cell)
  3. Row ordering preserved (Samples column)
  4. Column structure correct (59 original + 50 child PCs + 3 maternal PCs)
  5. Child PC values match source spreadsheet (via Inf_APrONID join)
  6. Unmatched child rows have NA for all 50 PCs
  7. No matched child row has NA PCs
  8. Maternal PC values match source spreadsheet (via MaternalID join)
  9. Unmatched maternal rows have NA for maternal PCs
 10. Child and maternal PCs are independent (not cross-contaminated)
 11. Join key logic spot-checks
 12. PC values are numeric where present
 13. Duplicate Inf_APrONID rows receive identical child PCs

Run from: the repository root
Usage:    python3 scripts/03_validate_harmonization.py

Input files (NOT modified):
  - data/Fulldata_raw.txt  (override via the RAW_DATA_FILE env var / config.sh)   (primary source)
  - data/APRON_PCs.xlsx                      (PCs source)
  - data/Fulldata_with_PCs_and_maternal_PCs.txt  (output to validate)

Output files:
  - results/validation_report.txt
"""

import csv
import os
import sys
from collections import defaultdict

try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl required. Install with: pip install openpyxl")
    sys.exit(1)

PRIMARY_FILE = os.environ.get("RAW_DATA_FILE", "data/Fulldata_raw.txt")
PCS_FILE = "data/APRON_PCs.xlsx"
OUTPUT_FILE = "data/Fulldata_with_PCs_and_maternal_PCs.txt"
REPORT_FILE = "results/validation_report.txt"
ENCODING = "latin-1"

errors = []
report_lines = []


def report(msg):
    print(msg)
    report_lines.append(msg)


def error(msg):
    errors.append(msg)
    report(f"  FAIL: {msg}")


def ok(msg):
    report(f"  OK: {msg}")


def load_tsv(filepath):
    with open(filepath, "r", encoding=ENCODING) as f:
        reader = csv.DictReader(f, delimiter="\t")
        return list(reader.fieldnames), list(reader)


def load_pcs():
    wb = openpyxl.load_workbook(PCS_FILE, read_only=True)
    ws = wb[wb.sheetnames[0]]
    headers = None
    data = {}
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            headers = list(row)
            continue
        data[row[0]] = {headers[j]: row[j] for j in range(1, len(headers))}
    wb.close()
    return headers[1:], data


def make_join_key(id_str):
    """Strip trailing visit suffix (-1 or -2)."""
    parts = id_str.rsplit("-", 1)
    if len(parts) == 2 and parts[1] in ("1", "2"):
        return parts[0]
    return id_str


def main():
    report("=" * 70)
    report("VALIDATION OF FULLY HARMONIZED DATASET")
    report("=" * 70)

    # Load everything independently
    report("\n[1] Loading all source files...")
    pri_fields, pri_rows = load_tsv(PRIMARY_FILE)
    out_fields, out_rows = load_tsv(OUTPUT_FILE)
    pc_cols, pc_data = load_pcs()
    ok(f"Primary: {len(pri_rows)} rows, {len(pri_fields)} cols")
    ok(f"Output:  {len(out_rows)} rows, {len(out_fields)} cols")
    ok(f"PCs:     {len(pc_data)} records, {len(pc_cols)} PCs")

    child_pc_cols = [f"PC{i}" for i in range(1, 51)]
    maternal_out_cols = ["maternal_PC1", "maternal_PC2", "maternal_PC3"]
    maternal_src_cols = ["PC1", "PC2", "PC3"]

    # Check 1: Row count
    report("\n[2] Row count...")
    if len(out_rows) != len(pri_rows):
        error(f"Row count: primary={len(pri_rows)}, output={len(out_rows)}")
    else:
        ok(f"Row count: {len(out_rows)}")

    # Check 2: Original data integrity (cell-by-cell)
    report("\n[3] Original data integrity (vs primary source)...")
    cells_checked = 0
    cells_wrong = 0
    for i in range(len(pri_rows)):
        for col in pri_fields:
            if pri_rows[i][col] != out_rows[i][col]:
                cells_wrong += 1
                if cells_wrong <= 3:
                    error(f"Row {i}, '{col}': primary='{pri_rows[i][col]}', "
                          f"output='{out_rows[i][col]}'")
            cells_checked += 1
    if cells_wrong == 0:
        ok(f"All {cells_checked:,} original cells match primary source")
    else:
        error(f"{cells_wrong} cells differ from primary source")

    # Check 3: Row ordering
    report("\n[4] Row ordering...")
    order_ok = True
    for i in range(len(pri_rows)):
        if pri_rows[i]["Samples"] != out_rows[i]["Samples"]:
            error(f"Row {i} order mismatch: "
                  f"primary='{pri_rows[i]['Samples']}', "
                  f"output='{out_rows[i]['Samples']}'")
            order_ok = False
            break
    if order_ok:
        ok("Row order preserved (verified via Samples column)")

    # Check 4: Column structure
    report("\n[5] Column structure...")
    expected_cols = pri_fields + child_pc_cols + maternal_out_cols
    if out_fields == expected_cols:
        ok(f"Column structure: {len(pri_fields)} original + "
           f"{len(child_pc_cols)} child PCs + "
           f"{len(maternal_out_cols)} maternal PCs = {len(expected_cols)}")
    else:
        if out_fields[:len(pri_fields)] != pri_fields:
            error("Original column names or order changed!")
        added = out_fields[len(pri_fields):]
        expected_added = child_pc_cols + maternal_out_cols
        if added != expected_added:
            error(f"Expected added cols: {len(expected_added)}, "
                  f"got {len(added)}")
            # Show first difference
            for j, (a, b) in enumerate(zip(expected_added, added)):
                if a != b:
                    error(f"  Position {j}: expected '{a}', got '{b}'")
                    break

    # Check 5: Child PC values vs source spreadsheet
    report("\n[6] Child PC values vs source spreadsheet...")
    child_checked = 0
    child_wrong = 0
    child_matched = 0

    for i in range(len(out_rows)):
        inf_id = out_rows[i]["Inf_APrONID"]
        jk = make_join_key(inf_id)
        if jk in pc_data:
            child_matched += 1
            for pc_col in child_pc_cols:
                expected = str(pc_data[jk][pc_col])
                actual = out_rows[i][pc_col]
                child_checked += 1
                if expected != actual:
                    child_wrong += 1
                    if child_wrong <= 3:
                        error(f"Row {i} ({inf_id}), {pc_col}: "
                              f"expected '{expected}', got '{actual}'")

    if child_wrong == 0:
        ok(f"All {child_checked:,} child PC values match source "
           f"({child_matched} rows x {len(child_pc_cols)} PCs)")
    else:
        error(f"{child_wrong} child PC cells wrong")

    # Check 6: Unmatched child rows have NA
    report("\n[7] Unmatched child rows have NA PCs...")
    child_unmatched = 0
    na_errors = 0
    for i in range(len(out_rows)):
        inf_id = out_rows[i]["Inf_APrONID"]
        jk = make_join_key(inf_id)
        if jk not in pc_data:
            child_unmatched += 1
            for pc_col in child_pc_cols:
                if out_rows[i][pc_col] != "NA":
                    na_errors += 1
                    if na_errors <= 3:
                        error(f"Row {i} ({inf_id}) unmatched but "
                              f"{pc_col}='{out_rows[i][pc_col]}'")
    if na_errors == 0:
        ok(f"All {child_unmatched} unmatched child rows have NA "
           f"for all {len(child_pc_cols)} PCs")

    # Check 7: No matched child row has NA PCs
    report("\n[8] No matched child row has NA PCs...")
    false_na = 0
    for i in range(len(out_rows)):
        inf_id = out_rows[i]["Inf_APrONID"]
        jk = make_join_key(inf_id)
        if jk in pc_data:
            for pc_col in child_pc_cols:
                if out_rows[i][pc_col] == "NA":
                    false_na += 1
                    if false_na <= 3:
                        error(f"Row {i} ({inf_id}) matched but "
                              f"{pc_col} is NA!")
    if false_na == 0:
        ok("No matched child rows have NA PCs")

    # Check 8: Maternal PC values vs source spreadsheet
    report("\n[9] Maternal PC values vs source spreadsheet...")
    mat_checked = 0
    mat_wrong = 0
    mat_matched = 0
    mat_unmatched = 0

    for i in range(len(out_rows)):
        mat_id = out_rows[i]["MaternalID"]
        jk = make_join_key(mat_id)
        if jk in pc_data:
            mat_matched += 1
            for src_col, out_col in zip(maternal_src_cols, maternal_out_cols):
                expected = str(pc_data[jk][src_col])
                actual = out_rows[i][out_col]
                mat_checked += 1
                if expected != actual:
                    mat_wrong += 1
                    if mat_wrong <= 3:
                        error(f"Row {i} ({mat_id}), {out_col}: "
                              f"expected '{expected}', got '{actual}'")
        else:
            mat_unmatched += 1
            for out_col in maternal_out_cols:
                if out_rows[i][out_col] != "NA":
                    mat_wrong += 1
                    error(f"Row {i} unmatched maternal but "
                          f"{out_col}='{out_rows[i][out_col]}'")

    if mat_wrong == 0:
        ok(f"All {mat_checked} maternal PC values verified against source")
        ok(f"All {mat_unmatched} unmatched maternal rows have NA")
    else:
        error(f"{mat_wrong} maternal PC issues found")

    # Check 9: Child vs maternal PC independence
    report("\n[10] Child vs maternal PC independence...")
    n_both = 0
    n_identical = 0
    for i in range(len(out_rows)):
        if (out_rows[i]["PC1"] != "NA"
                and out_rows[i]["maternal_PC1"] != "NA"):
            n_both += 1
            if out_rows[i]["PC1"] == out_rows[i]["maternal_PC1"]:
                n_identical += 1
    if n_identical == 0:
        ok(f"Child and maternal PC1 differ in all {n_both} rows "
           f"with both present (no cross-contamination)")
    else:
        error(f"{n_identical}/{n_both} rows have identical "
              f"child and maternal PC1!")

    # Check 10: Join key spot-checks
    report("\n[11] Join key logic spot-checks...")
    spot_checks = [
        ("C0123-10-2", "C0123-10"),
        ("C0262-10-2", "C0262-10"),
        ("C1160-10-1", "C1160-10"),
        ("C0123-01-2", "C0123-01"),
        ("C5200-10-1", "C5200-10"),
    ]
    all_spot_ok = True
    for test_id, expected_key in spot_checks:
        actual_key = make_join_key(test_id)
        if actual_key != expected_key:
            error(f"Join key for '{test_id}': "
                  f"expected '{expected_key}', got '{actual_key}'")
            all_spot_ok = False
    if all_spot_ok:
        ok(f"All {len(spot_checks)} join key spot-checks pass")

    # Check 11: PC values are numeric
    report("\n[12] PC values are numeric (not corrupted)...")
    non_numeric = 0
    for i in range(len(out_rows)):
        for pc_col in child_pc_cols[:5] + maternal_out_cols:
            val = out_rows[i][pc_col]
            if val == "NA":
                continue
            try:
                float(val)
            except ValueError:
                non_numeric += 1
                if non_numeric <= 3:
                    error(f"Row {i}, {pc_col}: non-numeric value '{val}'")
    if non_numeric == 0:
        ok("All non-NA PC values are valid numeric "
           "(checked child PC1-5 + maternal PCs, all rows)")

    # Check 12: Duplicate Inf_APrONID rows get same child PCs
    report("\n[13] Duplicate Inf_APrONID rows receive identical PCs...")
    id_to_rows = defaultdict(list)
    for i, r in enumerate(out_rows):
        id_to_rows[r["Inf_APrONID"]].append(i)
    dup_groups = {k: v for k, v in id_to_rows.items() if len(v) > 1}
    dup_ok = True
    for inf_id, row_indices in dup_groups.items():
        for pc_col in child_pc_cols:
            vals = set(out_rows[idx][pc_col] for idx in row_indices)
            if len(vals) > 1:
                error(f"Duplicate {inf_id}: inconsistent "
                      f"{pc_col} values: {vals}")
                dup_ok = False
    if dup_ok:
        ok(f"All {len(dup_groups)} duplicate ID groups have identical PCs")

    # Imputation-relevant cross-reference
    report("\n[14] Imputation-relevant cross-reference...")
    missing_child_has_maternal = 0
    missing_child_no_maternal = 0
    has_outcome_count = 0
    for i in range(len(out_rows)):
        likert = out_rows[i].get("Final7PointLikertScale", "").strip()
        if not likert:
            continue
        has_outcome_count += 1
        if out_rows[i]["PC1"] == "NA":
            if out_rows[i]["maternal_PC1"] != "NA":
                missing_child_has_maternal += 1
            else:
                missing_child_no_maternal += 1
    ok(f"Rows with outcome data: {has_outcome_count}")
    ok(f"Rows with outcome + missing child PCs + available maternal PCs: "
       f"{missing_child_has_maternal}")
    ok(f"Rows with outcome + missing child PCs + NO maternal PCs: "
       f"{missing_child_no_maternal}")

    # Summary
    report("\n" + "=" * 70)
    if errors:
        report(f"VALIDATION FAILED: {len(errors)} error(s)")
        for e in errors:
            report(f"  - {e}")
    else:
        report("VALIDATION PASSED: All checks passed, zero errors.")
    report("=" * 70)

    with open(REPORT_FILE, "w") as f:
        f.write("\n".join(report_lines) + "\n")
    print(f"Report: {REPORT_FILE}")

    return len(errors) == 0


if __name__ == "__main__":
    success = main()
    sys.exit(0 if success else 1)
