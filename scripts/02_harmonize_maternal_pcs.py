#!/usr/bin/env python3
"""
Step 2: Add maternal ancestry PCs to the harmonized dataset.

Joins maternal PCs from APRON_PCs.xlsx to the already-harmonized dataset
(Fulldata_with_PCs.txt) using MaternalID as the join key. The maternal PCs are
intended as AUXILIARY variables for MICE imputation only — they help impute
missing child PCs but are NOT included in the downstream analysis model.

Join strategy:
  - MaternalID in primary data: "C0123-01-2" (family-role-visit)
  - FID in PCs file:            "C0123-01"   (family-role, no visit suffix)
  - Strip trailing visit suffix (-1 or -2) from MaternalID to create join key.

Run from: the repository root
Usage:    python3 scripts/02_harmonize_maternal_pcs.py

Input files (NOT modified):
  - data/Fulldata_with_PCs.txt          (from step 01)
  - data/APRON_PCs.xlsx

Output files:
  - data/Fulldata_with_PCs_and_maternal_PCs.txt
  - results/harmonization_maternal_report.txt
"""

import csv
import sys

try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl required. Install with: pip install openpyxl")
    sys.exit(1)

INPUT_FILE = "data/Fulldata_with_PCs.txt"
PCS_FILE = "data/APRON_PCs.xlsx"
OUTPUT_FILE = "data/Fulldata_with_PCs_and_maternal_PCs.txt"
REPORT_FILE = "results/harmonization_maternal_report.txt"
ENCODING = "latin-1"

# Only include first 3 maternal PCs (auxiliary for imputation, not for analysis)
N_MATERNAL_PCS = 3


def read_input(filepath):
    with open(filepath, "r", encoding=ENCODING) as f:
        reader = csv.DictReader(f, delimiter="\t")
        fieldnames = list(reader.fieldnames)
        rows = list(reader)
    return fieldnames, rows


def read_pcs(filepath):
    wb = openpyxl.load_workbook(filepath, read_only=True)
    ws = wb[wb.sheetnames[0]]
    headers = None
    data = {}
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            headers = list(row)
            continue
        fid = row[0]
        data[fid] = {headers[j]: row[j] for j in range(1, len(headers))}
    wb.close()
    return headers[1:], data


def make_join_key(maternal_id):
    """Strip trailing visit suffix (-1 or -2) from MaternalID.

    Example: 'C0123-01-2' -> 'C0123-01'
    """
    parts = maternal_id.rsplit("-", 1)
    if len(parts) == 2 and parts[1] in ("1", "2"):
        return parts[0]
    return maternal_id


def main():
    report_lines = []

    def report(msg):
        print(msg)
        report_lines.append(msg)

    report("=" * 70)
    report("STEP 2: MATERNAL PC HARMONIZATION")
    report("=" * 70)
    report("")

    # Load data
    in_fields, in_rows = read_input(INPUT_FILE)
    report(f"Input: {len(in_rows)} rows, {len(in_fields)} cols")
    report(f"File: {INPUT_FILE}")

    pc_col_names, pc_data = read_pcs(PCS_FILE)
    report(f"PCs source: {len(pc_data)} records, {len(pc_col_names)} PCs")
    report(f"File: {PCS_FILE}")

    maternal_pc_cols = [f"PC{i}" for i in range(1, N_MATERNAL_PCS + 1)]
    maternal_out_cols = [f"maternal_{pc}" for pc in maternal_pc_cols]
    report(f"Maternal PCs to add: {maternal_out_cols}")
    report("")

    # Column collision check
    collisions = [c for c in maternal_out_cols if c in in_fields]
    if collisions:
        report(f"ERROR: Columns already exist: {collisions}")
        sys.exit(1)
    else:
        report(f"No column name collisions: OK")
    report("")

    # Maternal ID analysis
    mat_ids = [r["MaternalID"] for r in in_rows]
    join_keys = [make_join_key(mid) for mid in mat_ids]

    non_parent = [(i, jk) for i, jk in enumerate(join_keys)
                  if not jk.endswith("-01")]
    if non_parent:
        report(f"WARNING: {len(non_parent)} join keys are NOT parent IDs (-01):")
        for idx, jk in non_parent[:5]:
            report(f"  Row {idx}: {mat_ids[idx]} -> {jk}")
    else:
        report(f"All join keys are parent IDs (-01 suffix): OK")
    report("")

    # Perform the join
    matched, unmatched = [], []
    for i in range(len(in_rows)):
        (matched if join_keys[i] in pc_data else unmatched).append(i)

    report(f"Matched: {len(matched)}/{len(in_rows)} "
           f"({100*len(matched)/len(in_rows):.1f}%)")
    report(f"Unmatched: {len(unmatched)}/{len(in_rows)}")

    if unmatched:
        report("")
        report("Unmatched rows (maternal PCs will be NA):")
        for idx in unmatched:
            row = in_rows[idx]
            jk = join_keys[idx]
            child_pc1 = row.get("PC1", "?")
            child_status = ("has child PCs" if child_pc1 != "NA"
                            else "NO child PCs either")
            report(f"  Row {idx}: Sample={row['Samples']}, "
                   f"MaternalID={row['MaternalID']}, "
                   f"join_key={jk}, {child_status}")
    report("")

    # Cross-reference child vs maternal PC availability
    report("Cross-reference: child vs maternal PC availability")
    n_both = n_child_only = n_maternal_only = n_neither = 0
    for i, row in enumerate(in_rows):
        has_child = row.get("PC1", "NA") != "NA"
        has_maternal = join_keys[i] in pc_data
        if has_child and has_maternal:
            n_both += 1
        elif has_child:
            n_child_only += 1
        elif has_maternal:
            n_maternal_only += 1
        else:
            n_neither += 1

    report(f"  Both child & maternal PCs:     {n_both}")
    report(f"  Child PCs only (no maternal):  {n_child_only}")
    report(f"  Maternal PCs only (no child):  {n_maternal_only}"
           f"  <- these benefit from maternal auxiliary")
    report(f"  Neither child nor maternal:    {n_neither}")
    report("")

    # Build output
    output_fieldnames = list(in_fields) + maternal_out_cols
    output_rows = []
    for i, row in enumerate(in_rows):
        new_row = dict(row)
        jk = join_keys[i]
        if jk in pc_data:
            for pc_col, out_col in zip(maternal_pc_cols, maternal_out_cols):
                new_row[out_col] = pc_data[jk].get(pc_col, "NA")
        else:
            for out_col in maternal_out_cols:
                new_row[out_col] = "NA"
        output_rows.append(new_row)

    with open(OUTPUT_FILE, "w", encoding=ENCODING, newline="") as f:
        writer = csv.DictWriter(f, fieldnames=output_fieldnames, delimiter="\t")
        writer.writeheader()
        writer.writerows(output_rows)

    report(f"Output: {OUTPUT_FILE}")
    report(f"  {len(output_rows)} rows, {len(output_fieldnames)} columns "
           f"({len(in_fields)} prior + {len(maternal_out_cols)} maternal PCs)")
    report("")

    # Validation
    report("Validation:")
    assert len(output_rows) == len(in_rows)
    report(f"  Row count preserved: OK")

    for i, row in enumerate(in_rows):
        for col in in_fields:
            assert output_rows[i][col] == row[col], \
                f"Column {col} modified in row {i}!"
    report(f"  All {len(in_fields)} prior columns unchanged: OK")

    for i in matched:
        jk = join_keys[i]
        for pc_col, out_col in zip(maternal_pc_cols, maternal_out_cols):
            expected = str(pc_data[jk][pc_col])
            actual = str(output_rows[i][out_col])
            assert expected == actual, \
                f"Row {i} {out_col}: expected {expected}, got {actual}"
    report(f"  Maternal PC values verified for {len(matched)} matched rows: OK")

    for i in unmatched:
        for out_col in maternal_out_cols:
            assert output_rows[i][out_col] == "NA"
    report(f"  Unmatched rows have NA maternal PCs: OK ({len(unmatched)} rows)")

    # Child vs maternal PC independence check
    n_same = n_compared = 0
    for i in matched:
        if output_rows[i]["PC1"] != "NA":
            n_compared += 1
            if output_rows[i]["PC1"] == output_rows[i]["maternal_PC1"]:
                n_same += 1
    if n_compared > 0:
        report(f"  Child PC1 vs maternal_PC1: {n_same}/{n_compared} identical "
               f"(expected: mostly different)")
    report("")

    report("STEP 2 COMPLETE")
    report(f"  Neither input file was modified.")
    report(f"  maternal_PC1-3 are AUXILIARY variables for MICE imputation only.")
    report(f"  They should NOT be included in the downstream brms analysis model.")

    with open(REPORT_FILE, "w") as f:
        f.write("\n".join(report_lines) + "\n")
    print(f"Report: {REPORT_FILE}")


if __name__ == "__main__":
    main()
